"""Peek at Excel file structure - columns and sample data."""
import openpyxl
import sys

FILE = r"C:\Users\ASUS\Documents\Список селлеров b2b парсер\docs\File\STM_Sellers_Full_Master_v3_contacts_partial (1).xlsx"

wb = openpyxl.load_workbook(FILE, read_only=True)
print(f"Sheets: {wb.sheetnames}")
print("=" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n### Sheet: {sheet_name}")
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 5:
            break
        rows.append(row)

    if not rows:
        print("  (empty)")
        continue

    # Print header
    header = rows[0]
    print(f"  Columns ({len(header)}): {list(header)}")
    
    # Print sample rows
    for idx, row in enumerate(rows[1:], 1):
        print(f"  Row {idx}: {list(row)}")
    
    # Count total rows
    count = 0
    for _ in ws.iter_rows(values_only=True):
        count += 1
    print(f"  Total rows: {count - 1} (excluding header)")

wb.close()
