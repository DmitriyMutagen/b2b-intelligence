import openpyxl
import os

file_path = r"C:\Users\ASUS\Documents\Список селлеров b2b парсер\docs\File\STM_Sellers_Full_Master_v3_contacts_partial (1).xlsx"

try:
    wb = openpyxl.load_workbook(file_path, read_only=True)
    sheet = wb.active
    print(f"Sheet Name: {sheet.title}")
    
    # Get headers (first row)
    headers = []
    for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
        headers = list(row)
        break
        
    print("Headers FOUND:", headers)
    
    # Get first 3 rows of data
    print("\nFirst 3 data rows:")
    for row in sheet.iter_rows(min_row=2, max_row=4, values_only=True):
        print(row)
        
except Exception as e:
    print(f"Error reading with openpyxl: {e}")
